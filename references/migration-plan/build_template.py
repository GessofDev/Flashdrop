# -*- coding: utf-8 -*-
"""
Genera la plantilla de estimación para la migración de FlashDrop Backend.
Estructura:
  - Hoja 1: Catálogo de tipos de tarea con rúbrica y descripciones estándar.
  - Hoja 2: Mi estimación (items pre-cargados del plan + lugar para que
            cada junior complete su detalle).
  - Hoja 3: Resumen (totales por servicio, fase y responsable).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = r"D:\desarrollo\2027\flashdrop_backend\estimacion_migracion_template.xlsx"

# ---- Estilos ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHEADER_FONT = Font(bold=True, size=10)
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


# =========================================================================
# Sheet 1: Catálogo de tipos de tarea
# =========================================================================
catalogo_headers = [
    "Tipo de tarea",
    "Descripción genérica",
    "Estimación sugerida (h)",
    "Criterio de done",
    "Tests asociados",
    "Labels Jira sugeridos",
]

catalogo_rows = [
    [
        "Crear endpoint interno GET",
        "Implementar endpoint HTTP bajo /api/internal/** que retorna datos en formato JSON según contrato definido en el plan (sección 9).",
        2,
        "Endpoint responde con JSON correcto, validado contra el contrato. Test de integración 200/404 verde.",
        "Test de endpoint interno + Test de contrato JSON",
        "endpoint-interno, fase-1",
    ],
    [
        "Crear endpoint interno POST/PATCH",
        "Implementar endpoint HTTP con body request que muta estado (crear/actualizar recurso) en el servicio dueño.",
        3,
        "Endpoint responde 201/200 con JSON correcto. Validación de body y manejo de errores según formato estándar (sección 10).",
        "Test de integración HTTP (happy path + error 409/422)",
        "endpoint-interno, fase-1",
    ],
    [
        "Reemplazar adapter SQL → HTTP",
        "Cambiar la implementación de un port (interface hexagonal) que hoy consulta Supabase REST directo, por un HTTP client que llama al endpoint interno del servicio dueño. La firma del port NO cambia.",
        4,
        "Adapter nuevo implementa el port. Tests del use case pasan. No quedan referencias al adapter viejo.",
        "Test unitario del use case + Test de integración del HTTP client (mock o WireMock)",
        "adapter, fase-2",
    ],
    [
        "Crear esquema SQL independiente",
        "Definir el DDL de las tablas propias del servicio en un script SQL versionado. Cada tabla es propiedad exclusiva del servicio.",
        1,
        "Script SQL ejecutable contra la BD del servicio. Sin referencias a tablas de otros servicios.",
        "—",
        "bd, fase-3",
    ],
    [
        "Crear script de seed",
        "Generar script que popula la BD del servicio con datos demo realistas para desarrollo y tests.",
        1,
        "Script ejecutable idempotente. Datos suficientes para probar el flujo end-to-end.",
        "—",
        "seed, fase-3",
    ],
    [
        "Configurar nueva conexión BD",
        "Actualizar application.yml / .env del servicio para apuntar a su propia base de datos (SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY propios).",
        0.5,
        "Servicio levanta sin error de conexión. Queries a tablas propias funcionan.",
        "—",
        "config, fase-3",
    ],
    [
        "Configurar URLs servicios internos",
        "Definir variables de entorno con las URLs de los servicios que este servicio consume (AUTH_SERVICE_URL, CATALOG_SERVICE_URL, etc.) en application.yml.",
        0.5,
        "Servicio resuelve las URLs de sus dependencias. HTTP clients inyectan correctamente.",
        "—",
        "config, fase-2",
    ],
    [
        "Configurar InternalApiKeyFilter",
        "Implementar filtro HTTP que valida el header X-Internal-Api-Key en todos los endpoints bajo /api/internal/**. Si el header no coincide, responde 403.",
        0.5,
        "Endpoints internos rechazan requests sin X-Internal-Api-Key válido. Endpoints públicos NO se ven afectados.",
        "Test del filtro (200 con key válida, 403 sin key)",
        "config, fase-1",
    ],
    [
        "Test unitario de dominio",
        "Test que cubre la lógica de un use case o entidad de dominio sin tocar infraestructura (sin Spring, sin BD, sin HTTP).",
        1,
        "Test pasa. Cubre caso feliz + casos de error relevantes (excepciones de dominio).",
        "—",
        "test",
    ],
    [
        "Test de integración HTTP client",
        "Test que valida que un HTTP client (hacia otro servicio) envía el request correcto, parsea la respuesta, y maneja errores del servicio dependiente (caído, timeout, 404).",
        2,
        "Test pasa con MockRestServiceServer o WireMock. Cubre happy path + al menos un escenario de fallo.",
        "—",
        "test",
    ],
    [
        "Test de endpoint interno",
        "Test que valida un endpoint interno propio: request, respuesta esperada, status code. Usa MockMvc.",
        1,
        "Test pasa. Cubre 200 con datos y al menos un caso de error (404 o 400).",
        "—",
        "test",
    ],
    [
        "Test de contrato JSON",
        "Test que verifica que el JSON de respuesta de un endpoint interno tiene EXACTAMENTE los campos que los servicios consumidores esperan (nombres, tipos, nullabilidad).",
        0.5,
        "Test asserts sobre los nombres y tipos de cada campo del JSON.",
        "—",
        "test, contrato",
    ],
]


# =========================================================================
# Sheet 2: Items pre-cargados
# =========================================================================
estimacion_headers = [
    "ID", "Fase", "Servicio", "Responsable", "Tipo de tarea",
    "Título (resumen Jira)", "Descripción estándar", "Descripción específica (completar)",
    "Estimación (h)", "Dependencias", "Tests requeridos", "Estado",
    "Jira Key", "URL Jira", "Labels Jira sugeridos",
    "Fecha inicio", "Fecha fin", "Horas reales",
]

items = []
_id_counter = [0]


def add(fase, servicio, responsable, tipo, titulo, desc_std, deps, tests, labels):
    _id_counter[0] += 1
    items.append([
        _id_counter[0], fase, servicio, responsable, tipo, titulo, desc_std,
        "",                # descripción específica — la completa el junior
        "",                # estimación — la completa el junior
        deps, tests, "Pendiente",
        "",                # Jira Key
        "",                # URL Jira
        labels,
        "",                # fecha inicio
        "",                # fecha fin
        "",                # horas reales
    ])


# ---- Nicolás — Auth Service ---------------------------------------------
add(1, "Auth", "Nicolás", "Crear endpoint interno GET",
    "Crear GET /api/internal/users/{userId}",
    "Retorna {id, name, lastName, email, phone}. 404 si no existe.",
    "—", "Test de endpoint interno + Test de contrato JSON",
    "auth-service, fase-1, endpoint-interno")

add(1, "Auth", "Nicolás", "Crear endpoint interno GET",
    "Crear GET /api/internal/users/{userId}/roles",
    "Retorna array de {id, name}. Array vacío si no tiene roles.",
    "—", "Test de endpoint interno",
    "auth-service, fase-1, endpoint-interno")

add(1, "Auth", "Nicolás", "Configurar InternalApiKeyFilter",
    "Configurar filtro X-Internal-Api-Key en Auth",
    "Filtro valida header en /api/internal/**. 403 si falta o no coincide.",
    "—", "Test del filtro",
    "auth-service, fase-1, config")

# Tests unitarios Auth
for titulo, desc in [
    ("RegisterUserTest — email válido", "Email válido crea user + login + asigna rol."),
    ("RegisterUserTest — email duplicado", "Lanza EmailAlreadyRegisteredException."),
    ("RegisterUserTest — password débil", "Lanza WeakPasswordException."),
    ("AuthenticateUserTest — credenciales válidas", "Retorna JWT + refresh token."),
    ("AuthenticateUserTest — credenciales inválidas", "Lanza InvalidCredentialsException."),
    ("RefreshTokenTest — token válido", "Genera nuevo JWT."),
    ("RefreshTokenTest — token expirado", "Lanza InvalidTokenException."),
]:
    add(1, "Auth", "Nicolás", "Test unitario de dominio", titulo, desc,
        "—", "—", "auth-service, fase-1, test")

# Tests integración Auth
for titulo, desc in [
    ("GET /api/internal/users/{id} — 200 con datos", "Usuario existente retorna 200 con name, lastName, email, phone."),
    ("GET /api/internal/users/{id} — 404", "Usuario inexistente retorna 404."),
    ("GET /api/internal/users/{id}/roles — lista", "Retorna roles asignados."),
    ("GET /api/internal/users/{id}/roles — lista vacía", "Array vacío si no tiene roles."),
]:
    add(1, "Auth", "Nicolás", "Test de endpoint interno", titulo, desc,
        "—", "—", "auth-service, fase-1, test")

add(1, "Auth", "Nicolás", "Test de contrato JSON",
    "Contrato JSON de /api/internal/users/{id}",
    "JSON contiene exactamente: id, name, lastName, email, phone.",
    "—", "—", "auth-service, fase-1, test, contrato")

# Fase 3 Auth
add(3, "Auth", "Nicolás", "Crear esquema SQL independiente",
    "Crear esquema SQL Auth (users, login, roles, user_has_roles, refresh_tokens)",
    "DDL de las 5 tablas propias del servicio.",
    "—", "—", "auth-service, fase-3, bd")

add(3, "Auth", "Nicolás", "Configurar nueva conexión BD",
    "Configurar SUPABASE_URL y KEY propios en Auth",
    "Apuntar a la BD independiente del servicio Auth.",
    "Esquema SQL creado", "—", "auth-service, fase-3, config")

add(3, "Auth", "Nicolás", "Crear script de seed",
    "Crear seed con datos demo Auth",
    "Usuarios demo con login, roles asignados, refresh tokens.",
    "Esquema SQL creado", "—", "auth-service, fase-3, seed")


# ---- Javier — Catalog Service -------------------------------------------
add(1, "Catalog", "Javier", "Crear endpoint interno GET",
    "Crear GET /api/internal/products?ids=...",
    "Retorna array de {id, restaurantId, name, description, price, image, isAvailable}. Vacío si IDs no existen.",
    "—", "Test de endpoint interno + Test de contrato JSON",
    "catalog-service, fase-1, endpoint-interno")

add(1, "Catalog", "Javier", "Crear endpoint interno GET",
    "Crear GET /api/internal/restaurants/{restaurantId}",
    "Retorna {id, userId, name, address}. 404 si no existe.",
    "—", "Test de endpoint interno + Test de contrato JSON",
    "catalog-service, fase-1, endpoint-interno")

add(1, "Catalog", "Javier", "Crear endpoint interno GET",
    "Crear GET /api/internal/restaurants?userId=...",
    "Mismo formato que /{id}. 404 si el usuario no tiene restaurante.",
    "—", "Test de endpoint interno",
    "catalog-service, fase-1, endpoint-interno")

add(1, "Catalog", "Javier", "Configurar InternalApiKeyFilter",
    "Configurar filtro X-Internal-Api-Key en Catalog",
    "Filtro valida header en /api/internal/**.",
    "—", "Test del filtro",
    "catalog-service, fase-1, config")

# Tests unitarios Catalog
for titulo, desc in [
    ("ListProductsTest — por categoría", "Retorna solo productos de la categoría."),
    ("ListProductsTest — por restaurante", "Retorna solo productos del restaurante."),
    ("CreateProductTest — precio negativo", "Lanza excepción de dominio."),
    ("CreateProductTest — categoría inexistente", "Lanza excepción."),
    ("GetProductsByIdsTest", "IDs válidos retornan productos; inexistentes se ignoran."),
    ("ListCategoriesTest", "Retorna todas las categorías ordenadas."),
    ("ListRestaurantsTest", "Retorna todos los restaurantes."),
]:
    add(1, "Catalog", "Javier", "Test unitario de dominio", titulo, desc,
        "—", "—", "catalog-service, fase-1, test")

# Tests integración Catalog
for titulo, desc in [
    ("GET /api/internal/products — 200 con productos", "Retorna productos con id, restaurantId, name, description, price, image, isAvailable."),
    ("GET /api/internal/products?ids=999 — lista vacía", "IDs inexistentes retornan array vacío."),
    ("GET /api/internal/restaurants/{id} — 200", "Retorna 200 con id, userId, name, address."),
    ("GET /api/internal/restaurants/{id} — 404", "Restaurante inexistente retorna 404."),
    ("GET /api/internal/restaurants?userId=...", "Retorna restaurante del usuario o 404."),
]:
    add(1, "Catalog", "Javier", "Test de endpoint interno", titulo, desc,
        "—", "—", "catalog-service, fase-1, test")

for titulo, desc in [
    ("Contrato JSON de /api/internal/products", "Campos: id, restaurantId, name, description, price, image, isAvailable."),
    ("Contrato JSON de /api/internal/restaurants/{id}", "Campos: id, userId, name, address."),
]:
    add(1, "Catalog", "Javier", "Test de contrato JSON", titulo, desc,
        "—", "—", "catalog-service, fase-1, test, contrato")

# Fase 3 Catalog
add(3, "Catalog", "Javier", "Crear esquema SQL independiente",
    "Crear esquema SQL Catalog (categories, products, restaurant)",
    "DDL de las 3 tablas propias del servicio.",
    "—", "—", "catalog-service, fase-3, bd")

add(3, "Catalog", "Javier", "Configurar nueva conexión BD",
    "Configurar SUPABASE_URL y KEY propios en Catalog",
    "Apuntar a la BD independiente del servicio Catalog.",
    "Esquema SQL creado", "—", "catalog-service, fase-3, config")

add(3, "Catalog", "Javier", "Crear script de seed",
    "Crear seed con datos demo Catalog",
    "Categorías, productos y restaurantes demo.",
    "Esquema SQL creado", "—", "catalog-service, fase-3, seed")


# ---- Felipe — Orders Service --------------------------------------------
add(1, "Orders", "Felipe", "Crear endpoint interno GET",
    "Crear GET /api/internal/orders?ids=...",
    "Retorna array de {id, clientId, restaurantId, deliveryId, status, address}. Vacío si IDs no existen.",
    "—", "Test de endpoint interno",
    "orders-service, fase-1, endpoint-interno")

add(1, "Orders", "Felipe", "Configurar InternalApiKeyFilter",
    "Configurar filtro X-Internal-Api-Key en Orders",
    "Filtro valida header en /api/internal/**.",
    "—", "Test del filtro",
    "orders-service, fase-1, config")

for titulo, desc in [
    ("GET /api/internal/orders?ids=1,2,3 — 200", "Retorna las órdenes existentes con sus campos."),
    ("GET /api/internal/orders?ids=999 — lista vacía", "IDs inexistentes retornan array vacío."),
]:
    add(1, "Orders", "Felipe", "Test de endpoint interno", titulo, desc,
        "—", "—", "orders-service, fase-1, test")

# Fase 2 Orders — adapters
add(2, "Orders", "Felipe", "Reemplazar adapter SQL → HTTP",
    "SupabaseRestCatalogAdapter → CatalogHttpClientAdapter",
    "Implementar CatalogPort vía HTTP client a /api/internal/products y /api/internal/restaurants.",
    "Endpoints internos de Catalog (Fase 1)", "Test de integración HTTP client",
    "orders-service, fase-2, adapter")

add(2, "Orders", "Felipe", "Reemplazar adapter SQL → HTTP",
    "Accesos a users en SupabaseRestDeliveryAdapter → AuthHttpClientAdapter",
    "Leer users vía /api/internal/users/{id} en vez de Supabase REST.",
    "Endpoints internos de Auth (Fase 1)", "Test de integración HTTP client",
    "orders-service, fase-2, adapter")

add(2, "Orders", "Felipe", "Reemplazar adapter SQL → HTTP",
    "Accesos a delivery → DeliveryHttpClientAdapter",
    "Leer delivery vía /api/internal/delivery-persons?userId=...",
    "Endpoints internos de Delivery (Fase 1)", "Test de integración HTTP client",
    "orders-service, fase-2, adapter")

add(2, "Orders", "Felipe", "Reemplazar adapter SQL → HTTP",
    "Escritura de delivery_routes → POST /api/internal/routes",
    "OrderRepositoryAdapter debe llamar a POST /api/internal/routes en vez de escribir directo.",
    "Endpoints internos de Delivery (Fase 1)", "Test de integración HTTP client",
    "orders-service, fase-2, adapter")

add(2, "Orders", "Felipe", "Configurar URLs servicios internos",
    "Configurar AUTH_SERVICE_URL, CATALOG_SERVICE_URL, DELIVERY_SERVICE_URL en Orders",
    "Definir URLs de servicios que Orders consume en application.yml.",
    "—", "—", "orders-service, fase-2, config")

# Tests unitarios Orders
for titulo, desc in [
    ("CreateOrderTest — productos válidos", "Calcula subtotal, delivery_fee y total correctamente."),
    ("CreateOrderTest — producto no disponible", "Lanza excepción de dominio."),
    ("CreateOrderTest — lista vacía", "Lanza excepción."),
    ("UpdateOrderStatusTest — transición válida", "Transición (ej: 'Nuevo pedido' → 'Preparando') se aplica."),
    ("UpdateOrderStatusTest — transición inválida", "Lanza excepción."),
    ("ClaimOrdersTest", "Asignar repartidor actualiza delivery_id y status."),
]:
    add(2, "Orders", "Felipe", "Test unitario de dominio", titulo, desc,
        "—", "—", "orders-service, fase-2, test")

# Tests integración HTTP Orders
for titulo, desc in [
    ("CatalogHttpClientAdapter — IDs válidos", "Retorna lista de productos con precio."),
    ("CatalogHttpClientAdapter — Catalog caído", "Lanza excepción manejable (no 500 genérico)."),
    ("AuthHttpClientAdapter — válido", "Retorna datos del usuario."),
    ("AuthHttpClientAdapter — usuario inexistente", "Retorna Optional.empty (no excepción)."),
    ("DeliveryHttpClientAdapter — info repartidor", "Retorna info del repartidor."),
    ("DeliveryHttpClientAdapter — crear ruta 201", "Creación de ruta vía Delivery API retorna 201."),
]:
    add(2, "Orders", "Felipe", "Test de integración HTTP client", titulo, desc,
        "Adapters HTTP creados", "—", "orders-service, fase-2, test")

# Fase 3 Orders
add(3, "Orders", "Felipe", "Crear esquema SQL independiente",
    "Crear esquema SQL Orders (orders, order_items, client)",
    "DDL de las 3 tablas propias del servicio.",
    "—", "—", "orders-service, fase-3, bd")

add(3, "Orders", "Felipe", "Configurar nueva conexión BD",
    "Configurar SUPABASE_URL y KEY propios en Orders",
    "Apuntar a la BD independiente del servicio Orders.",
    "Esquema SQL creado", "—", "orders-service, fase-3, config")

add(3, "Orders", "Felipe", "Crear script de seed",
    "Crear seed con datos demo Orders",
    "Órdenes demo con items y clientes.",
    "Esquema SQL creado", "—", "orders-service, fase-3, seed")


# ---- Sebastián — Delivery Service ---------------------------------------
add(1, "Delivery", "Sebastián", "Crear endpoint interno GET",
    "Crear GET /api/internal/delivery-persons?userId=...",
    "Retorna {id, userId, vehicle}. 404 si no existe.",
    "—", "Test de endpoint interno",
    "delivery-service, fase-1, endpoint-interno")

add(1, "Delivery", "Sebastián", "Crear endpoint interno POST/PATCH",
    "Crear POST /api/internal/routes",
    "Body: orderId, pickupAddress, deliveryAddress, distanceKm, estimatedMinutes, status. 201 con ruta. 409 si ya existe.",
    "—", "Test de endpoint interno",
    "delivery-service, fase-1, endpoint-interno")

add(1, "Delivery", "Sebastián", "Crear endpoint interno POST/PATCH",
    "Crear PATCH /api/internal/routes/{orderId}/status",
    "Body: {status}. 200 con ruta actualizada. 404 si no existe.",
    "—", "Test de endpoint interno",
    "delivery-service, fase-1, endpoint-interno")

add(1, "Delivery", "Sebastián", "Configurar InternalApiKeyFilter",
    "Configurar filtro X-Internal-Api-Key en Delivery",
    "Filtro valida header en /api/internal/**.",
    "—", "Test del filtro",
    "delivery-service, fase-1, config")

for titulo, desc in [
    ("GET /api/internal/delivery-persons — 200", "Repartidor existente retorna 200 con id, userId, vehicle."),
    ("GET /api/internal/delivery-persons — 404", "Usuario sin perfil de repartidor retorna 404."),
    ("POST /api/internal/routes — 201", "Body válido crea ruta y retorna 201."),
    ("POST /api/internal/routes — 409 (duplicado)", "Lanza RouteAlreadyAssignedException."),
    ("PATCH /api/internal/routes — 200", "Actualiza status y retorna 200."),
    ("PATCH /api/internal/routes — 404", "Ruta inexistente retorna 404."),
]:
    add(1, "Delivery", "Sebastián", "Test de endpoint interno", titulo, desc,
        "—", "—", "delivery-service, fase-1, test")

# Fase 2 Delivery
add(2, "Delivery", "Sebastián", "Reemplazar adapter SQL → HTTP",
    "OrderServiceClientAdapter → HTTP clients a Orders y Catalog",
    "Llamar a /api/internal/orders y /api/internal/restaurants en vez de Supabase REST.",
    "Endpoints internos de Orders y Catalog (Fase 1)", "Test de integración HTTP client",
    "delivery-service, fase-2, adapter")

for titulo, desc in [
    ("ClaimDeliveryTest — mismo restaurante", "Repartidor válido reclama órdenes del mismo restaurante."),
    ("ClaimDeliveryTest — distintos restaurantes", "Lanza excepción."),
    ("ClaimDeliveryTest — repartidor inexistente", "Lanza DeliveryPersonNotFoundException."),
    ("UpdateRouteStatusTest — transición válida", "Se aplica correctamente."),
    ("UpdateRouteStatusTest — ruta inexistente", "Lanza RouteNotFoundException."),
    ("ListRoutesTest", "Retorna rutas del repartidor."),
]:
    add(2, "Delivery", "Sebastián", "Test unitario de dominio", titulo, desc,
        "—", "—", "delivery-service, fase-2, test")

for titulo, desc in [
    ("OrdersHttpClientAdapter — órdenes con dirección", "Retorna órdenes con dirección y restaurant_id."),
    ("OrdersHttpClientAdapter — Orders caído", "Lanza excepción manejable."),
    ("CatalogHttpClientAdapter — nombre y dirección", "Retorna nombre y dirección del restaurante."),
    ("CatalogHttpClientAdapter — fallback", "Restaurante inexistente retorna fallback (no rompe el flujo)."),
]:
    add(2, "Delivery", "Sebastián", "Test de integración HTTP client", titulo, desc,
        "Adapters HTTP creados", "—", "delivery-service, fase-2, test")

# Fase 3 Delivery
add(3, "Delivery", "Sebastián", "Crear esquema SQL independiente",
    "Crear esquema SQL Delivery (delivery, delivery_routes)",
    "DDL de las 2 tablas propias del servicio.",
    "—", "—", "delivery-service, fase-3, bd")

add(3, "Delivery", "Sebastián", "Configurar nueva conexión BD",
    "Configurar SUPABASE_URL y KEY propios en Delivery",
    "Apuntar a la BD independiente del servicio Delivery.",
    "Esquema SQL creado", "—", "delivery-service, fase-3, config")

add(3, "Delivery", "Sebastián", "Crear script de seed",
    "Crear seed con datos demo Delivery",
    "Repartidores y rutas demo.",
    "Esquema SQL creado", "—", "delivery-service, fase-3, seed")


# ---- Fase 4 — compartida -----------------------------------------------
add(4, "Todos", "Todos", "Test de integración HTTP client",
    "Test e2e del flujo completo (Fase 4)",
    "Registro → login → ver catálogo → crear orden → asignar repartidor → actualizar estado de ruta. Verificar que no haya queries fallidas por tablas faltantes.",
    "Fases 1+2+3 completas", "—",
    "fase-4, e2e, shared")


# =========================================================================
# Construcción del workbook
# =========================================================================
wb = Workbook()

# ---- Sheet 1: Catálogo ----
ws = wb.active
ws.title = "Catálogo"

for col, h in enumerate(catalogo_headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER

for r, row in enumerate(catalogo_rows, 2):
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.alignment = LEFT if c != 3 else CENTER
        cell.border = BORDER

widths = [32, 60, 18, 50, 35, 32]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Dropdown para estimación sugerida
dv = DataValidation(type="list", formula1='"0.5,1,2,4,8,16"', allow_blank=True)
dv.error = "Estimación debe ser una de: 0.5, 1, 2, 4, 8, 16"
dv.errorTitle = "Valor inválido"
ws.add_data_validation(dv)
dv.add(f"C2:C{len(catalogo_rows) + 1}")

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 28


# ---- Sheet 2: Mi estimación ----
ws2 = wb.create_sheet("Mi estimación")

for col, h in enumerate(estimacion_headers, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER

n_items = len(items)
last_row = n_items + 1

for r, row in enumerate(items, 2):
    for c, v in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.alignment = LEFT if c in (7, 8) else CENTER
        cell.border = BORDER

# Anchos
widths2 = [6, 8, 12, 14, 32, 50, 50, 40, 14, 28, 28, 14, 12, 40, 30, 14, 14, 14]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Data validations
dv_est = DataValidation(type="list", formula1='"0.5,1,2,4,8,16"', allow_blank=True)
dv_est.error = "Estimación debe ser una de: 0.5, 1, 2, 4, 8, 16"
dv_est.errorTitle = "Valor inválido"
ws2.add_data_validation(dv_est)
dv_est.add(f"I2:I{last_row}")

dv_estado = DataValidation(type="list", formula1='"Pendiente,En progreso,Completado"', allow_blank=True)
ws2.add_data_validation(dv_estado)
dv_estado.add(f"L2:L{last_row}")

dv_serv = DataValidation(type="list", formula1='"Auth,Catalog,Orders,Delivery,Todos"', allow_blank=True)
ws2.add_data_validation(dv_serv)
dv_serv.add(f"C2:C{last_row}")

dv_fase = DataValidation(type="list", formula1='"1,2,3,4"', allow_blank=True)
ws2.add_data_validation(dv_fase)
dv_fase.add(f"B2:B{last_row}")

# Tipo de tarea — referenciar el catálogo
tipos_list = ",".join(t.replace('"', '""') for t in [r[0] for r in catalogo_rows])
dv_tipo = DataValidation(type="list", formula1=f'"{tipos_list}"', allow_blank=True)
dv_tipo.error = "Tipo debe existir en el Catálogo"
dv_tipo.errorTitle = "Tipo inválido"
ws2.add_data_validation(dv_tipo)
dv_tipo.add(f"E2:E{last_row}")

# VLOOKUP para auto-completar Descripción estándar (col G) desde Catálogo
for r in range(2, last_row + 1):
    ws2.cell(row=r, column=7).value = f'=IFERROR(VLOOKUP(E{r},Catálogo!A:B,2,FALSE),"")'

ws2.freeze_panes = "A2"
ws2.row_dimensions[1].height = 32


# ---- Sheet 3: Resumen ----
ws3 = wb.create_sheet("Resumen")

ws3.cell(row=1, column=1, value="Resumen de estimación — Migración FlashDrop Backend").font = Font(bold=True, size=14, color="1F4E78")
ws3.merge_cells("A1:D1")

# Por servicio
ws3.cell(row=3, column=1, value="Por servicio").font = Font(bold=True, size=12, color="1F4E78")

for col, h in enumerate(["Servicio", "Items", "Horas estimadas", "Horas reales"], 1):
    c = ws3.cell(row=4, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER

servicios = ["Auth", "Catalog", "Orders", "Delivery", "Todos"]
for i, s in enumerate(servicios, 5):
    ws3.cell(row=i, column=1, value=s).border = BORDER
    ws3.cell(row=i, column=2, value=f"=COUNTIF('Mi estimación'!C:C,A{i})").border = BORDER
    ws3.cell(row=i, column=3, value=f"=SUMIFS('Mi estimación'!I:I,'Mi estimación'!C:C,A{i})").border = BORDER
    ws3.cell(row=i, column=4, value=f"=SUMIFS('Mi estimación'!R:R,'Mi estimación'!C:C,A{i})").border = BORDER
    for c in (2, 3, 4):
        ws3.cell(row=i, column=c).alignment = CENTER

total_row = 5 + len(servicios)
ws3.cell(row=total_row, column=1, value="Total")
for c in range(1, 5):
    cell = ws3.cell(row=total_row, column=c)
    cell.fill = SUBHEADER_FILL
    cell.font = Font(bold=True)
    cell.border = BORDER
ws3.cell(row=total_row, column=2, value=f"=SUM(B5:B{total_row-1})")
ws3.cell(row=total_row, column=3, value=f"=SUM(C5:C{total_row-1})")
ws3.cell(row=total_row, column=4, value=f"=SUM(D5:D{total_row-1})")
for c in (2, 3, 4):
    ws3.cell(row=total_row, column=c).alignment = CENTER

# Por fase
fase_start = total_row + 3
ws3.cell(row=fase_start, column=1, value="Por fase").font = Font(bold=True, size=12, color="1F4E78")

for col, h in enumerate(["Fase", "Items", "Horas estimadas", "Horas reales"], 1):
    c = ws3.cell(row=fase_start + 1, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER

for i, f in enumerate([1, 2, 3, 4], fase_start + 2):
    ws3.cell(row=i, column=1, value=f).border = BORDER
    ws3.cell(row=i, column=2, value=f"=COUNTIF('Mi estimación'!B:B,A{i})").border = BORDER
    ws3.cell(row=i, column=3, value=f"=SUMIFS('Mi estimación'!I:I,'Mi estimación'!B:B,A{i})").border = BORDER
    ws3.cell(row=i, column=4, value=f"=SUMIFS('Mi estimación'!R:R,'Mi estimación'!B:B,A{i})").border = BORDER
    for c in (1, 2, 3, 4):
        ws3.cell(row=i, column=c).alignment = CENTER

# Por responsable
resp_start = fase_start + 7
ws3.cell(row=resp_start, column=1, value="Por responsable").font = Font(bold=True, size=12, color="1F4E78")

for col, h in enumerate(["Responsable", "Items", "Horas estimadas", "Horas reales"], 1):
    c = ws3.cell(row=resp_start + 1, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER

responsables = ["Nicolás", "Javier", "Felipe", "Sebastián", "Todos"]
for i, r in enumerate(responsables, resp_start + 2):
    ws3.cell(row=i, column=1, value=r).border = BORDER
    ws3.cell(row=i, column=2, value=f"=COUNTIF('Mi estimación'!D:D,A{i})").border = BORDER
    ws3.cell(row=i, column=3, value=f"=SUMIFS('Mi estimación'!I:I,'Mi estimación'!D:D,A{i})").border = BORDER
    ws3.cell(row=i, column=4, value=f"=SUMIFS('Mi estimación'!R:R,'Mi estimación'!D:D,A{i})").border = BORDER
    for c in (1, 2, 3, 4):
        ws3.cell(row=i, column=c).alignment = CENTER

# Por tipo
tipo_start = resp_start + 8
ws3.cell(row=tipo_start, column=1, value="Por tipo de tarea").font = Font(bold=True, size=12, color="1F4E78")

for col, h in enumerate(["Tipo", "Items", "Horas estimadas", "Horas reales"], 1):
    c = ws3.cell(row=tipo_start + 1, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER

tipos = [r[0] for r in catalogo_rows]
for i, t in enumerate(tipos, tipo_start + 2):
    ws3.cell(row=i, column=1, value=t).border = BORDER
    ws3.cell(row=i, column=2, value=f"=COUNTIF('Mi estimación'!E:E,A{i})").border = BORDER
    ws3.cell(row=i, column=3, value=f"=SUMIFS('Mi estimación'!I:I,'Mi estimación'!E:E,A{i})").border = BORDER
    ws3.cell(row=i, column=4, value=f"=SUMIFS('Mi estimación'!R:R,'Mi estimación'!E:E,A{i})").border = BORDER
    for c in (1, 2, 3, 4):
        ws3.cell(row=i, column=c).alignment = CENTER

# Anchos
ws3.column_dimensions["A"].width = 38
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 20
ws3.column_dimensions["D"].width = 20

ws3.freeze_panes = "A4"

# ---- Guardar ----
wb.save(OUTPUT)
print(f"OK: {OUTPUT}")
print(f"Items pre-cargados: {n_items}")
print(f"Tipos de tarea en catálogo: {len(catalogo_rows)}")
